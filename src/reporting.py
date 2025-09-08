import pandas as pd
from datetime import datetime
from pathlib import Path
from .config import REPORTS_PATH, DATA_PROCESSED_PATH, RULE_BASED_REPORT_PATH, LLM_REPORT_PATH, FINAL_OUTPUT_PATH
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def generate_quality_report(check_results, report_path=RULE_BASED_REPORT_PATH):
    """Generate comprehensive data quality report."""
    with open(report_path, 'w') as f:
        f.write("Data Quality Assessment Report\n")
        f.write(f"Generated: {datetime.now()}\n")
        f.write("=" * 50 + "\n\n")
        
        # Data type issues
        if check_results['dtype_issues']:
            f.write("** Data Type Issues:\n")
            for col, info in check_results['dtype_issues'].items():
                f.write(f" - {col}: expected {info['expected']}, got {info['actual']}\n")
        else:
            f.write("** All data types are correct.\n")
        f.write("\n")
        
        # Missing values
        f.write("** Missing Values:\n")
        for col, count in check_results['missing_summary'].items():
            f.write(f" - {col}: {count} missing\n")
        f.write("\n")
        
        # Duplicates
        dup_count = len(check_results['duplicates'])
        f.write(f"** Duplicate Records: {dup_count}\n\n")
        
        # Volatility
        vol_count = check_results['volatility_flags']['company_name'].nunique()
        f.write(f"** Companies with High Volatility: {vol_count}\n\n")
        
        # --- Date Correction Summary ---
        f.write("** Date Format Analysis:\n")
        f.write(f" - Records corrected: {check_results.get('date_corrections_count', 0)}\n")
        f.write(f" - Records with remaining format issues: {check_results.get('remaining_date_issues', 0)}\n")
        f.write("\n")

def generate_llm_report(llm_results, report_path=LLM_REPORT_PATH, values_standardized=0):
    """Generate LLM analysis report."""
    with open(report_path, 'w') as f:
        f.write("LLM Anomaly Detection Report\n")
        f.write(f"Generated: {datetime.now()}\n")
        f.write("=" * 50 + "\n\n")
        
        # Adding standardization info
        f.write("** Missing Value Standardization:\n")
        f.write(f" - Missing values standardized to 'N/A': {values_standardized}\n")
        f.write("\n")
        
        f.write("** LLM Analysis Results:\n")
        for company, result in llm_results.items():
            f.write(f"Company: {company}\n")
            f.write(f"Verdict: {result.get('verdict', 'N/A')}\n")
            f.write(f"Confidence: {result.get('confidence', 'N/A')}\n")
            f.write(f"Explanation: {result.get('explanation', 'N/A')}\n")
            f.write("-" * 40 + "\n\n")

def save_to_excel(df, filepath=FINAL_OUTPUT_PATH):
    """Save final results to Excel with professional formatting and sorting."""
    # --- Identify LLM-processed vs non-LLM-processed records ---
    # Check if any of the LLM columns have meaningful data (not "N/A")
    llm_processed_mask = (
        (df['llm_verdict'] != "N/A") | 
        (df['llm_explanation'] != "N/A") | 
        (df['llm_confidence'] != "N/A")
    )
    
    llm_processed_count = llm_processed_mask.sum()
    non_llm_count = len(df) - llm_processed_count
    
    if non_llm_count > 0:
        print(f"**<>  NOTE: Found {non_llm_count} records without LLM analysis!")
        print("   These records will be placed after LLM-analyzed records.")
    
    # --- Identify records with missing company names ---
    missing_company_mask = df['company_name'].isna() | (df['company_name'] == "N/A") | (df['company_name'] == "")
    missing_company_count = missing_company_mask.sum()
    
    if missing_company_count > 0:
        print(f"**<>  WARNING: Found {missing_company_count} records with missing company names!")
    
    # --- Create sorted copy ---
    # Separate LLM-processed and non-LLM-processed records
    llm_df = df[llm_processed_mask].copy()
    non_llm_df = df[~llm_processed_mask].copy()
    
    # Sort LLM-processed records: first by company name A-Z, then by year (newest first)
    if not llm_df.empty:
        llm_df = llm_df.sort_values(
            by=['company_name', 'year'], 
            ascending=[True, False]
        )
    
    # Sort non-LLM records 
    if not non_llm_df.empty:
        non_llm_df = non_llm_df.sort_values(
            by=['company_name', 'year'], 
            ascending=[True, False]
        )
    
    # Combine them: LLM-processed first, then non-LLM
    sorted_df = pd.concat([llm_df, non_llm_df], ignore_index=True)
    
    # --- Save with formatting ---
    sorted_df.to_excel(filepath, index=False)
    
    # Apply professional formatting
    wb = load_workbook(filepath)
    ws = wb.active

    # Wrap text for LLM explanation column
    if "llm_explanation" in df.columns:
        col_idx = df.columns.get_loc("llm_explanation") + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Freeze header row for easy scrolling
    ws.freeze_panes = "A2"
    
    # Highlight ONLY rows with missing company names (not all non-LLM rows)
    if missing_company_count > 0:
        from openpyxl.styles import PatternFill
        red_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
        
        company_name_col = sorted_df.columns.get_loc('company_name') + 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            company_cell = ws.cell(row=row_idx, column=company_name_col)
            if company_cell.value in [None, "", "N/A"]:
                for cell in row:
                    cell.fill = red_fill
    
    wb.save(filepath)
    
    # --- Print summary ---
    llm_companies = llm_df['company_name'].nunique() if not llm_df.empty else 0
    print(f"✓ Output sorted: {llm_companies} companies with LLM analysis, {non_llm_count} records without LLM analysis")
    print(f"✓ LLM-analyzed companies sorted alphabetically, then by year (newest first)")
    print(f"✓ Only records with missing company names are highlighted in red")
    
    return filepath